#!/usr/bin/env python3
"""Import the legacy CRM Excel (Statusy_z_CRM.xlsx) into the NocoDB CRM base.

What it does per legacy row:
  - creates a lead (selects normalized case-insensitively against real column options)
  - creates a company from 'Organizacja' (find-or-create by name) and links it
  - converts the 12 legacy date columns into: milestone fields on the lead
    (received_at, offer_sent_at, contract_sent_at, closed_at) + done `meetings`
    (badanie potrzeb, DEMO, omówienie oferty)
  - turns 'Data planowanego działania' + 'Planowane działanie' into an open task
  - computes enquiry_no (1..n per contact e-mail, ordered by received date)
  - logs an `activities` entry with the FULL raw row in payload (nothing is lost,
    even for columns this script does not map)
  - is idempotent: rows whose legacy_id already exists in NocoDB are skipped

Required extra fields on `leads` (add before running):
  legacy_id (Number), received_at (DateTime)

Usage:
  pip install requests openpyxl
  python3 import_legacy_excel.py --dry-run   # ALWAYS first: mapping report, no writes
  python3 import_legacy_excel.py             # real import
"""
import argparse
import json
import re
import sys
import unicodedata
from collections import defaultdict
from datetime import date, datetime, time, timedelta

import openpyxl
import requests

# ----------------------------------------------------------------- CONFIG
CONFIG = {
    "url": "https://noco.example.com",
    "token": "PASTE_NOCODB_API_TOKEN",
    "base_id": "PASTE_BASE_ID",
    "xlsx": "Statusy_z_CRM.xlsx",
    # Legacy 'Handlowiec' value -> NocoDB user e-mail:
    "owner_map": {"przemek": "przemek@example.com"},
    "default_owner": "przemek@example.com",
}
TABLES = ["companies", "leads", "meetings", "tasks", "activities"]

# Legacy 'Etap' -> new stage option. Extend after the dry-run report.
STAGE_MAP = {
    "nowy": "new",
    "umówione discovery": "discovery_scheduled",
    "po discovery": "discovery_done",
    "badanie potrzeb": "discovery_done",
    "audyt": "audit",
    "rekomendacja": "recommendation",
    "oferta wysłana": "offer_sent",
    "omówienie oferty": "offer_discussed",
    "umowa wysłana": "contract_sent",
    "umowa podpisana": "contract_signed",
    "utracona": "lost",
    "archiwum": "archived",
}
SOURCE_MAP = {"google": "google", "polecenie": "polecenie", "linkedin": "linkedin"}
CHANNEL_MAP = {"bookings": "bookings", "formularz": "formularz",
               "e-mail": "email", "mail": "email", "telefon": "telefon"}
PUBLIC_DOMAINS = {"gmail.com", "googlemail.com", "wp.pl", "o2.pl", "onet.pl", "op.pl",
                  "interia.pl", "icloud.com", "outlook.com", "hotmail.com", "live.com",
                  "proton.me", "protonmail.com", "yahoo.com", "gazeta.pl", "poczta.fm"}

# ----------------------------------------------------------------- API helpers
S = requests.Session()
S.headers.update({"xc-token": CONFIG["token"], "Content-Type": "application/json"})
BASE = CONFIG["url"].rstrip("/")


def api(method, path, **kw):
    r = S.request(method, f"{BASE}{path}", **kw)
    if not r.ok:
        sys.exit(f"API error {r.status_code} on {method} {path}: {r.text[:500]}")
    return r.json() if r.text else {}


def resolve_meta():
    tables, links, selects = {}, {}, {}
    for t in api("GET", f"/api/v2/meta/bases/{CONFIG['base_id']}/tables").get("list", []):
        title = t["title"].strip().lower()
        if title in TABLES:
            tables[title] = t["id"]
    missing = [t for t in TABLES if t not in tables]
    if missing:
        sys.exit(f"Tables not found: {missing}")
    for title, tid in tables.items():
        links[title], selects[title] = {}, {}
        for col in api("GET", f"/api/v2/meta/tables/{tid}").get("columns", []):
            ct = col["title"].strip().lower()
            if col.get("uidt") in ("Links", "LinkToAnotherRecord"):
                links[title][ct] = col["id"]
            elif col.get("uidt") in ("SingleSelect", "MultiSelect"):
                selects[title][ct] = [o["title"] for o in
                                      (col.get("colOptions") or {}).get("options", [])]
    return tables, links, selects


UNMAPPED = defaultdict(set)   # (table, field) -> set of skipped values


def norm_selects(table, record):
    out = {}
    for k, v in record.items():
        opts = SEL.get(table, {}).get(k.strip().lower())
        if opts is not None and isinstance(v, str):
            match = next((o for o in opts if o.lower() == v.lower()), None)
            if match is None:
                UNMAPPED[(table, k)].add(v)
                continue
            v = match
        out[k] = v
    return out


def create(table, record):
    rec = norm_selects(table, record)
    if DRY:
        return None
    res = api("POST", f"/api/v2/tables/{TBL[table]}/records", json=rec)
    return res.get("Id") or res.get("id")


def link(table, field, record_id, target_ids):
    if DRY or record_id is None:
        return
    fid = LNK[table].get(field)
    if not fid:
        print(f"  ! link field '{field}' missing on '{table}'")
        return
    ids = target_ids if isinstance(target_ids, list) else [target_ids]
    ids = [i for i in ids if i is not None]
    if ids:
        api("POST", f"/api/v2/tables/{TBL[table]}/links/{fid}/records/{record_id}",
            json=[{"Id": i} for i in ids])


# ----------------------------------------------------------------- Excel helpers
def clean_header(h):
    return re.sub(r"\s+", " ", str(h or "")).strip()


def as_date(v):
    """Cell -> ISO date string. Handles datetime, date and raw Excel serials."""
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    if isinstance(v, (int, float)):   # Excel serial fallback
        return (date(1899, 12, 30) + timedelta(days=int(v))).isoformat()
    try:
        return datetime.fromisoformat(str(v)).date().isoformat()
    except ValueError:
        return None


def as_dt(dv, tv):
    """Combine a date cell and a time cell into an ISO datetime."""
    d = as_date(dv)
    if not d:
        return None
    if isinstance(tv, time):
        return f"{d}T{tv.strftime('%H:%M:%S')}"
    if isinstance(tv, (int, float)):  # time as fraction of a day
        secs = int(round(float(tv) % 1 * 86400))
        return f"{d}T{secs // 3600:02d}:{secs % 3600 // 60:02d}:00"
    return f"{d}T00:00:00"


def fold(s):
    return unicodedata.normalize("NFD", str(s or "")).encode("ascii", "ignore") \
        .decode().lower().strip()


def json_safe(v):
    if isinstance(v, (datetime, date, time)):
        return v.isoformat()
    return v


# ----------------------------------------------------------------- main
ap = argparse.ArgumentParser()
ap.add_argument("--dry-run", action="store_true")
DRY = ap.parse_args().dry_run

print("Resolving base metadata..." if not DRY else "DRY RUN - nothing will be written.")
TBL, LNK, SEL = ({}, {}, {}) if False else resolve_meta()

existing_legacy = set()
if not DRY:
    res = api("GET", f"/api/v2/tables/{TBL['leads']}/records"
                     "?fields=legacy_id&limit=1000")
    existing_legacy = {r.get("legacy_id") for r in res.get("list", []) if r.get("legacy_id")}
    print(f"Already imported legacy_ids: {len(existing_legacy)}")

wb = openpyxl.load_workbook(CONFIG["xlsx"], data_only=True)
ws = wb[wb.sheetnames[0]]
rows = list(ws.iter_rows(values_only=True))
headers = [clean_header(h) for h in rows[0]]
data = [dict(zip(headers, r)) for r in rows[1:] if any(v not in (None, "") for v in r)]
print(f"Legacy rows found: {len(data)}")

# enquiry_no: number leads 1..n per e-mail (fallback: folded name), by received date
groups = defaultdict(list)
for row in data:
    key = fold(row.get("E.mail")) or f"name:{fold(row.get('Nazwa klienta'))}"
    groups[key].append(row)
enquiry_no = {}
for key, rws in groups.items():
    rws.sort(key=lambda r: as_date(r.get("Data wpłynięcia")) or "9999")
    for i, r in enumerate(rws, 1):
        enquiry_no[id(r)] = i

company_cache = {}   # folded name -> company Id
stats = defaultdict(int)

for row in data:
    lid = row.get("ID")
    if lid in existing_legacy:
        stats["skipped_existing"] += 1
        continue

    owner = CONFIG["owner_map"].get(fold(row.get("Handlowiec")), CONFIG["default_owner"])
    b2x = "B2B" if "b2b" in fold(row.get("B2B / B2C")) else "B2C"
    email = str(row.get("E.mail") or "").strip().lower()

    # --- state & stage ------------------------------------------------------
    signed = as_date(row.get("Data podpisania umowy"))
    lost_d = as_date(row.get("Data utracenia"))
    arch_d = as_date(row.get("Data archiwizacji"))
    stan = fold(row.get("Stan"))
    if stan == "otwarta":
        state = "open"
    elif signed:
        state = "won"
    elif lost_d or row.get("Powód utraty szansy"):
        state = "lost"
    else:
        state = "archived"
    stage = STAGE_MAP.get(fold(row.get("Etap")))
    if row.get("Etap") and not stage:
        UNMAPPED[("leads", "stage(Etap)")].add(str(row.get("Etap")))

    # --- company ------------------------------------------------------------
    company_id = None
    org = clean_header(row.get("Organizacja"))
    if org:
        ck = fold(org)
        if ck not in company_cache:
            domain = email.split("@")[1] if "@" in email else ""
            company_cache[ck] = create("companies", {
                "name": org,
                "domains": domain if domain and domain not in PUBLIC_DOMAINS else "",
                "folder_url": str(row.get("Folder klienta") or "")})
            stats["companies"] += 1
        company_id = company_cache[ck]

    # --- lead -----------------------------------------------------------------
    notes = str(row.get("Notatki") or "")
    if row.get("Folder klienta") and not org:
        notes += f"\nFolder klienta: {row.get('Folder klienta')}"
    if row.get("Spr. ID"):
        notes += f"\nLegacy Spr. ID: {row.get('Spr. ID')}"
    lead = {
        "legacy_id": lid,
        "contact_name": clean_header(row.get("Osoba kontaktowa")) if b2x == "B2B"
                        and row.get("Osoba kontaktowa")
                        else clean_header(row.get("Nazwa klienta")) or "Unknown",
        "contact_email": email, "contact_phone": str(row.get("Nr telefonu") or ""),
        "type": b2x, "owner": owner,
        "source": SOURCE_MAP.get(fold(row.get("Źródło")), str(row.get("Źródło") or "")) or None,
        "contact_channel": CHANNEL_MAP.get(fold(row.get("Forma kontaktu")),
                                           str(row.get("Forma kontaktu") or "")) or None,
        "qualification": str(row.get("Kwalifikacja lead'a") or "") or None,
        "disqualify_note": str(row.get("Powód braku kwalifikacji lead'a") or "") or None,
        "industry": str(row.get("Branża") or "") or None,
        "stage": stage, "state": state,
        "loss_note": str(row.get("Powód utraty szansy") or "") or None,
        "value": row.get("Szansa sprzedaży Wartość"),
        "label": str(row.get("Szansa sprzedaży . Etykieta") or "") or None,
        "notes": notes.strip() or None,
        "enquiry_no": enquiry_no[id(row)],
        "received_at": as_dt(row.get("Data wpłynięcia"), row.get("Godzina wpłynięcia")),
        "offer_sent_at": as_date(row.get("Data wysłania oferty")),
        "contract_sent_at": as_date(row.get("Data wysłania umowy")),
        "closed_at": signed or lost_d or arch_d,
        "offer_prep_status": "none", "company_match_status": "none",
        "duplicate_check": "none",
    }
    lead = {k: v for k, v in lead.items() if v not in (None, "")}
    lead_id = create("leads", lead)
    stats["leads"] += 1
    link("leads", "company", lead_id, company_id)

    # --- legacy dates that were meetings -------------------------------------
    who = lead.get("contact_name", "")
    for col, mtype, title in [
            ("Data badania potrzeb", "needs_analysis", "Badanie potrzeb"),
            ("Data DEMO", "demo", "DEMO"),
            ("Data omówienia oferty", "offer_discussion", "Omówienie oferty")]:
        md = as_date(row.get(col))
        if md:
            mid = create("meetings", {
                "title": f"{title} — {who}", "type": mtype,
                "starts_at": f"{md}T00:00:00", "owner": owner,
                "status": "done", "processing_status": "none",
                "notes": "Imported from legacy Excel."})
            link("meetings", "lead", mid, lead_id)
            stats["meetings"] += 1

    # --- planned action -> open task ---------------------------------------------
    plan_d = as_date(row.get("Data planowanego działania"))
    plan_t = str(row.get("Planowane działanie") or "").strip()
    if plan_d or plan_t:
        tid = create("tasks", {
            "title": plan_t or f"Zaplanowane działanie: {who}",
            "status": "todo", "priority": "normal", "assignee": owner,
            "due_date": plan_d, "created_by_flow": "import",
            "description": f"lead:{lead_id}\nImported from legacy Excel."})
        link("tasks", "lead", tid, lead_id)
        stats["tasks"] += 1

    # --- traceability activity with the FULL raw row --------------------------------
    raw = {k: json_safe(v) for k, v in row.items() if v not in (None, "")}
    aid = create("activities", {
        "summary": f"Lead imported from legacy Excel (ID {lid}): {who}",
        "type": "lead_created", "flow": "import", "triggered_by": "system",
        "payload": json.dumps(raw, ensure_ascii=False)})
    link("activities", "lead", aid, lead_id)

# ----------------------------------------------------------------- report
print("\n===== REPORT =====")
for k, v in sorted(stats.items()):
    print(f"{k}: {v}")
if UNMAPPED:
    print("\nUNMAPPED VALUES (fix the *_MAP dicts or add options to the select, then re-run):")
    for (table, field), vals in UNMAPPED.items():
        print(f"  {table}.{field}: {sorted(vals)}")
elif DRY:
    print("\nAll values map cleanly - run again without --dry-run.")
if DRY:
    print("(dry run - nothing was written)")
